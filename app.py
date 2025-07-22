import os
import json
import base64
import tempfile
import time
from typing import List, Dict, Any, Tuple, Optional, BinaryIO, Union
from pathlib import Path

import streamlit as st
import pandas as pd
import pdfplumber
import docx
from PIL import Image
import pytesseract
import requests
from dotenv import load_dotenv

# Import the document classifier and parser
from document_classifier import detect_document_type, get_extraction_hints
from document_parser import parse_uploaded_file, StructuredElement

# Load environment variables
load_dotenv()

# App title and description
st.set_page_config(page_title="pdf2dataset", layout="wide")

# ----------------------------
# Header (title + tagline)
# ----------------------------
st.markdown(
    """
    <style>
        /* Reduce default Streamlit top padding */
        div.block-container {
            padding-top: 2rem; /* Increased to avoid clipping */
        }

        /* Header styling */
        .pdf2dataset-header {
            display: flex;
            align-items: baseline;
            column-gap: 0.75rem;
            font-family: "Helvetica Neue", Helvetica, Arial, sans-serif;
        }

        .pdf2dataset-header h1 {
            font-size: 2.4rem;
            font-weight: 800;
            margin: 0;
            color: #ffffff; /* Ensure visibility on dark background */
        }

        .pdf2dataset-header span.tagline {
            font-size: 1.25rem;
            font-weight: 400;
            color: #cccccc; /* Light grey for contrast */
        }
    </style>

    <div class="pdf2dataset-header">
        <h1>pdf2dataset</h1>
    </div>
    """,
    unsafe_allow_html=True,
)

# ------------------------------------------------------------------
# Utility: stringify nested dicts/lists so Streamlit tables render nicely
# ------------------------------------------------------------------

def sanitize_df(df: pd.DataFrame) -> pd.DataFrame:
    """Return a copy where any dict/list cells are JSON-stringified for display."""
    return df.map(lambda x: json.dumps(x, ensure_ascii=False) if isinstance(x, (dict, list)) else x)


def extract_text_from_pdf(file: BinaryIO) -> Tuple[str, List[str], Dict[str, Any], List[Dict[str, Any]]]:
    """
    Extract text from a PDF file using pdfplumber.
    
    Args:
        file: The uploaded PDF file
        
    Returns:
        Tuple containing extracted text, list of failed pages, document info, and structured elements
    """
    text = ""
    failed_pages = []
    document_info = {}
    structured_content = []
    
    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as temp_file:
        temp_file.write(file.getvalue())
        temp_path = temp_file.name
    
    try:
        # Use the document parser to extract structured content
        output_dir = os.path.join(tempfile.gettempdir(), f"pdf2dataset_{int(time.time())}")
        os.makedirs(output_dir, exist_ok=True)
        
        structured_content = parse_uploaded_file(file, file.name, output_dir)
        
        # Extract text from structured content
        for item in structured_content:
            if item.get("type") == "text" or item.get("content_type") == "text":
                text += item.get("content", "") + "\n"
            elif item.get("type") == "table" or item.get("content_type") == "table":
                text += item.get("content", "") + "\n\n"
            elif (item.get("type") == "metadata" or item.get("content_type") == "metadata") and item.get("metadata", {}).get("document_type"):
                document_info = item.get("metadata", {})
            elif item.get("type") == "question" or item.get("content_type") == "question":
                text += item.get("content", "") + "\n"
                # Also include any options as text
                if "options" in item:
                    for letter, option_text in item["options"].items():
                        text += f"{letter}) {option_text}\n"
        
        # If no document type detected yet, use traditional method for backward compatibility
        if not document_info:
            with pdfplumber.open(temp_path) as pdf:
                total_pages = len(pdf.pages)
                progress_bar = st.progress(0)
                
                for i, page in enumerate(pdf.pages):
                    try:
                        page_text = page.extract_text() or ""
                        text += page_text + "\n"
                    except Exception:
                        failed_pages.append(str(i + 1))
                    
                    progress_bar.progress((i + 1) / total_pages)
                
                # Detect document type from extracted text
                if text:
                    doc_info = detect_document_type(text)
                    document_info = {
                        "document_type": doc_info.get("document_type", "unknown"),
                        "confidence": doc_info.get("confidence", 0),
                        "extraction_hints": get_extraction_hints(doc_info.get("document_type", "unknown"))
                    }
    except Exception as e:
        st.error(f"Failed to process PDF: {e}")
    finally:
        # Clean up temp file
        try:
            Path(temp_path).unlink(missing_ok=True)
        except:
            pass
    
    return text, failed_pages, document_info, structured_content


def extract_text_from_docx(file: BinaryIO) -> Tuple[str, Dict[str, Any], List[Dict[str, Any]]]:
    """
    Extract text from a DOCX file using python-docx.
    
    Args:
        file: The uploaded DOCX file
        
    Returns:
        Tuple containing extracted text, document info, and structured elements
    """
    text = ""
    document_info = {}
    structured_content = []
    
    with tempfile.NamedTemporaryFile(delete=False, suffix=".docx") as temp_file:
        temp_file.write(file.getvalue())
        temp_path = temp_file.name
    
    try:
        # Use the document parser to extract structured content
        output_dir = os.path.join(tempfile.gettempdir(), f"pdf2dataset_{int(time.time())}")
        os.makedirs(output_dir, exist_ok=True)
        
        structured_content = parse_uploaded_file(file, file.name, output_dir)
        
        # Extract text from structured content
        for item in structured_content:
            if item.get("type") == "text" or item.get("content_type") == "text":
                text += item.get("content", "") + "\n"
            elif item.get("type") == "table" or item.get("content_type") == "table":
                text += item.get("content", "") + "\n\n"
            elif (item.get("type") == "metadata" or item.get("content_type") == "metadata") and item.get("metadata", {}).get("document_type"):
                document_info = item.get("metadata", {})
            elif item.get("type") == "question" or item.get("content_type") == "question":
                text += item.get("content", "") + "\n"
                # Also include any options as text
                if "options" in item:
                    for letter, option_text in item["options"].items():
                        text += f"{letter}) {option_text}\n"
        
        # If no document type detected yet, use traditional method for backward compatibility
        if not text or not document_info:
            doc = docx.Document(temp_path)
            text = "\n".join([paragraph.text for paragraph in doc.paragraphs])
            
            # Detect document type from extracted text
            if text:
                doc_info = detect_document_type(text)
                document_info = {
                    "document_type": doc_info.get("document_type", "unknown"),
                    "confidence": doc_info.get("confidence", 0),
                    "extraction_hints": get_extraction_hints(doc_info.get("document_type", "unknown"))
                }
    except Exception as e:
        st.error(f"Failed to process DOCX: {e}")
        return "", {}, []
    finally:
        # Clean up temp file
        try:
            Path(temp_path).unlink(missing_ok=True)
        except:
            pass
    
    return text, document_info, structured_content


def extract_text_from_image(file: BinaryIO) -> Tuple[str, Dict[str, Any], List[Dict[str, Any]]]:
    """
    Extract text from an image file using pytesseract.
    
    Args:
        file: The uploaded image file
        
    Returns:
        Tuple containing extracted text, document info, and structured elements
    """
    text = ""
    document_info = {}
    structured_content = []
    
    try:
        # Use the document parser to extract structured content
        output_dir = os.path.join(tempfile.gettempdir(), f"pdf2dataset_{int(time.time())}")
        os.makedirs(output_dir, exist_ok=True)
        
        structured_content = parse_uploaded_file(file, file.name, output_dir)
        
        # Extract text from structured content
        for item in structured_content:
            if item.get("type") == "text" or item.get("content_type") == "text":
                text += item.get("content", "") + "\n"
            elif (item.get("type") == "metadata" or item.get("content_type") == "metadata") and item.get("metadata", {}).get("document_type"):
                document_info = item.get("metadata", {})
            elif item.get("type") == "question" or item.get("content_type") == "question":
                text += item.get("content", "") + "\n"
                # Also include any options as text
                if "options" in item:
                    for letter, option_text in item["options"].items():
                        text += f"{letter}) {option_text}\n"
        
        # If no document type detected yet, use traditional method for backward compatibility
        if not text:
            image = Image.open(file)
            text = pytesseract.image_to_string(image)
            
            # Detect document type from extracted text
            if text:
                doc_info = detect_document_type(text)
                document_info = {
                    "document_type": doc_info.get("document_type", "unknown"),
                    "confidence": doc_info.get("confidence", 0),
                    "extraction_hints": get_extraction_hints(doc_info.get("document_type", "unknown"))
                }
    except Exception as e:
        st.error(f"Failed to process image: {e}")
        return "", {}, []
    
    return text, document_info, structured_content


def call_llm(prompt: str, text: str, document_info: Dict[str, Any] = None) -> List[Dict[str, Any]]:
    """
    Send text and prompt to an LLM API and get structured data in response.
    
    Args:
        prompt: User's natural language prompt describing desired dataset
        text: Extracted text from documents
        document_info: Optional document classification information
        
    Returns:
        List of dictionaries containing structured data
    """
    api_key = os.getenv("LLM_API_KEY")
    api_provider = os.getenv("LLM_PROVIDER", "openrouter").lower()
    model = os.getenv("LLM_MODEL", "")
    
    if not api_key:
        st.error("LLM API key not found. Please add it to your .env file.")
        st.info("Set LLM_API_KEY=your_key and optionally LLM_PROVIDER=openrouter|openai")
        return []
    
    # Enhanced system prompt to instruct the LLM
    system_prompt = """
    You are a JSON extraction API.
    
    TASK:
    Convert the provided unstructured text into a valid **JSON array** that fulfils the user's extraction request.
    
    OUTPUT RULES:
    1. Output **ONLY** a valid JSON array – absolutely no markdown fences, commentary, or extra text.
    2. Preserve every mathematical symbol or special formatting by wrapping the **entire value** in back-ticks, e.g.
         "question": "`f'(x) = 2x`".
    3. Use exactly the keys requested by the user. If none are given, infer sensible, consistent field names.
    4. Be comprehensive – extract **all** matching records.
    5. If nothing matches, output an empty array `[]`.
    """
    
    # Enhanced full prompt with clearer instructions and document type info
    full_prompt = f"""
    # EXTRACTION REQUEST
    {prompt}
    
    # DOCUMENT TYPE INFORMATION
    """
    
    # Add document type information if available
    if document_info and document_info.get("document_type") != "unknown":
        doc_type = document_info.get("document_type", "")
        extraction_hints = document_info.get("extraction_hints", {})
        
        if doc_type:
            full_prompt += f"Document Type: {doc_type}\n\n"
            
            if extraction_hints:
                full_prompt += "Expected Structure:\n"
                if extraction_hints.get("fields"):
                    fields_str = ", ".join(extraction_hints["fields"])
                    full_prompt += f"- Expected fields: {fields_str}\n"
                if extraction_hints.get("structure"):
                    full_prompt += f"- Structure: {extraction_hints['structure']}\n"
                full_prompt += "\n"
    
    full_prompt += f"""
    # EXTRACTED TEXT TO PROCESS
    ```
    {text[:100000]}
    ```
    
    # INSTRUCTIONS
    1. Analyze the text above to find ALL instances matching the extraction request
    2. Create appropriate column names based on the data and request
    3. Extract ALL matching data entries (complete rows)
    4. For tables or lists, preserve the relationships between columns
    5. Return ONLY a valid JSON array with properly formatted data
    
    Remember to:
    - Include ALL relevant records that match the criteria
    - Use consistent column names across all records
    - Use the most appropriate data type for each value
    - Do NOT return any explanations, only the JSON array
    """
    
    # Configure API based on provider
    if api_provider == "openai":
        url = "https://api.openai.com/v1/chat/completions"
        headers = {
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json"
        }
        # Use provided model or default to gpt-3.5-turbo (cheaper than gpt-4)
        model_name = model if model else "gpt-3.5-turbo"
        payload = {
            "model": model_name,
            "messages": [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": full_prompt}
            ],
            "response_format": {"type": "json_object"},
            "temperature": 0  # Deterministic output
        }
    else:  # Default to OpenRouter
        url = "https://openrouter.ai/api/v1/chat/completions"
        headers = {
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
            "HTTP-Referer": "https://pdf2dataset-app.local", # Required for OpenRouter
            "X-Title": "pdf2dataset App"  # Optional for OpenRouter
        }
        # Use provided model or default to a free model
        if model:
            model_name = model
        else:
            # Free models on OpenRouter (as of current date)
            model_name = "mistralai/mistral-7b-instruct:free"
            
        payload = {
            "model": model_name,
            "messages": [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": full_prompt}
            ],
            "response_format": {"type": "json_object"},
            "temperature": 0  # Deterministic output
        }
    
    # Attempt API call with retry
    max_retries = 3
    retry_delay = 2  # seconds
    tried_without_rf = False
    
    for attempt in range(max_retries):
        try:
            with st.spinner(f"Calling AI service (attempt {attempt+1}/{max_retries})..."):
                response = requests.post(url, headers=headers, json=payload)
                response.raise_for_status()
                try:
                    response_data = response.json()
                except json.JSONDecodeError:
                    # Provider returned non-JSON (e.g., Cloudflare HTML or plain text). Show it.
                    st.error("LLM provider returned non-JSON payload.")
                    st.code(response.text[:2000] + ("..." if len(response.text) > 2000 else ""))
                    # Do not retry endlessly – break and return empty list so app continues.
                    return []

                if "choices" not in response_data or not response_data["choices"]:
                    # Unexpected response format; surface error details if available
                    err_detail = response_data.get("error") or response_data
                    st.error("Unexpected LLM API response – no choices returned")
                    st.json(err_detail)
                    return []

                content = response_data["choices"][0]["message"].get("content", "")

                # ------------------------------------------------------------------
                # Robust JSON extraction: some models wrap the JSON with commentary
                # ------------------------------------------------------------------

                def _extract_json_block(raw: str) -> str:
                    """Return the first well-balanced JSON object/array found in *raw*.

                    1. Strips markdown fences and smart quotes.
                    2. Locates the first '[' or '{' then walks forward keeping
                       a depth counter until the matching closing bracket.
                    """
                    # Normalise text
                    txt = raw.strip()
                    # Remove markdown code fences if present
                    if txt.startswith("```"):
                        parts = txt.split("```")
                        if len(parts) >= 2:
                            txt = parts[1].strip()

                    # Replace smart quotes that break JSON parsers
                    txt = (
                        txt.replace("“", "\"")
                        .replace("”", "\"")
                        .replace("‘", "'")
                        .replace("’", "'")
                    )

                    # Quick path
                    try:
                        json.loads(txt)
                        return txt
                    except Exception:
                        pass

                    import re

                    m = re.search(r"[\[{]", txt)
                    if not m:
                        return ""

                    start = m.start()
                    opening = txt[start]
                    closing = "]" if opening == "[" else "}"

                    depth = 0
                    in_str = False
                    escape = False

                    for i in range(start, len(txt)):
                        ch = txt[i]

                        if in_str:
                            if escape:
                                escape = False
                            elif ch == "\\":
                                escape = True
                            elif ch == '"':
                                in_str = False
                            continue

                        if ch == '"':
                            in_str = True
                        elif ch == opening:
                            depth += 1
                        elif ch == closing:
                            depth -= 1
                            if depth == 0:
                                candidate = txt[start : i + 1]
                                try:
                                    json.loads(candidate)
                                    return candidate
                                except Exception:
                                    return ""

                    return ""

                json_block = _extract_json_block(content)

                if not json_block:
                    st.error("Failed to locate JSON in LLM response")
                    with st.expander("LLM raw reply"):
                        st.code(content)
                    st.info("Try simplifying your request or using a different model.")
                    return []

                # Try to parse the extracted JSON block
                try:
                    result = json.loads(json_block)
                    # If the result is a dict with a records key, use that (some APIs wrap the result)
                    if isinstance(result, dict) and "records" in result:
                        return result["records"]
                    # If the result is already a list of dicts, use it directly
                    elif isinstance(result, list) and len(result) > 0 and isinstance(result[0], dict):
                        return result
                    # If it's a dictionary but not in the expected format, try to extract data
                    elif isinstance(result, dict):
                        # Look for any key that might contain the records
                        for key, value in result.items():
                            if isinstance(value, list) and len(value) > 0 and isinstance(value[0], dict):
                                return value
                        # If we can't find a list, create a single record from the dict itself
                        return [result]
                    else:
                        st.error(f"Unexpected response format from LLM: {type(result)}")
                        st.info("Try rephrasing your request to be more specific about the data structure you want.")
                        return []
                except json.JSONDecodeError:
                    st.error("Failed to parse extracted JSON block from LLM response")
                    st.info("Try simplifying your request or using a different model.")
                    return []
                
            # If successful, break out of retry loop
            break
            
        except requests.exceptions.HTTPError as e:
            if e.response.status_code == 429:  # Too Many Requests
                if attempt < max_retries - 1:  # Don't sleep on the last attempt
                    st.warning(f"Rate limit exceeded. Retrying in {retry_delay} seconds...")
                    time.sleep(retry_delay)
                    retry_delay *= 2  # Exponential backoff
                else:
                    st.error(f"Rate limit exceeded. Please try again later or use a different API provider.")
                    st.info("You can set LLM_PROVIDER=openai in your .env file to use OpenAI directly.")
                    return []
            elif e.response.status_code == 402:  # Payment Required
                st.error("Payment required for this model. Please check your account balance or use a free model.")
                if api_provider == "openrouter":
                    st.info("Try adding this to your .env file: LLM_MODEL=mistralai/mistral-7b-instruct:free")
                else:
                    st.info("Try using a model you have access to with LLM_MODEL=gpt-3.5-turbo")
                return []
            else:
                # Some models (esp. on OpenRouter) don't accept response_format.
                # Retry once without that field.
                if (not tried_without_rf and "response_format" in payload):
                    tried_without_rf = True
                    payload.pop("response_format", None)
                    st.warning("Retrying LLM call without enforced JSON response_format...")
                    continue
                st.error(f"HTTP error calling LLM API: {e}")
                try:
                    st.json(e.response.json())
                except Exception:
                    st.write(e.response.text)
                return []
        except Exception as e:
            st.error(f"Error calling LLM API: {e}")
            return []
    
    return []  # Return empty list if all retries failed


def get_table_download_link(df: pd.DataFrame, file_format: str) -> str:
    """
    Generate a download link for the dataframe in the specified format.
    
    Args:
        df: Pandas DataFrame to download
        file_format: Format of the file to download ('csv', 'excel', or 'json')
        
    Returns:
        HTML string containing download link
    """
    if file_format == "csv":
        csv = df.to_csv(index=False)
        b64 = base64.b64encode(csv.encode()).decode()
        href = f'<a href="data:file/csv;base64,{b64}" download="dataset.csv">Download CSV</a>'
        return href
    elif file_format == "excel":
        # Use BytesIO instead of temporary file to avoid file locking issues
        import io
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
        
        # Process the dataframe to detect and preserve numeric formats
        df_processed = df.copy()
        numeric_formats = {}
        
        # Detect numeric columns and their appropriate formats
        for col in df.columns:
            # Check if column contains numeric data
            if pd.api.types.is_numeric_dtype(df[col]):
                # Try to determine if it's a percentage, currency, etc.
                sample_values = df[col].dropna().head(10).tolist()
                if len(sample_values) > 0:
                    # Default format
                    numeric_formats[col] = 'General'
                    
                    # Check for percentages (values between 0-1 or with % in column name)
                    if ('percent' in col.lower() or 'pct' in col.lower() or '%' in col.lower()):
                        numeric_formats[col] = '0.00%'
                    # Check for currency
                    elif ('price' in col.lower() or 'cost' in col.lower() or 'amount' in col.lower() 
                          or 'revenue' in col.lower() or '$' in col.lower() or 'dollar' in col.lower()
                          or 'eur' in col.lower() or 'gbp' in col.lower()):
                        numeric_formats[col] = '$#,##0.00'
                    # Check for general numbers needing commas
                    elif any(abs(val) >= 1000 for val in sample_values if val is not None):
                        numeric_formats[col] = '#,##0.00'
                    # General decimal format
                    else:
                        numeric_formats[col] = '0.00'
            
            # Check for date-like columns
            elif pd.api.types.is_datetime64_any_dtype(df[col]):
                numeric_formats[col] = 'yyyy-mm-dd'
        
        buffer = io.BytesIO()
        
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df_processed.to_excel(writer, index=False, sheet_name='Data')
            
            # Access the worksheet to apply formatting
            worksheet = writer.sheets['Data']
            
            # Format the header row
            for col_num, value in enumerate(df_processed.columns.values, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="E0E0E0")
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    bottom=Side(border_style='thin'),
                    top=Side(border_style='thin'),
                    left=Side(border_style='thin'),
                    right=Side(border_style='thin')
                )
            
            # Auto-adjust columns width and apply numeric formats
            for i, column in enumerate(df_processed.columns):
                col_letter = get_column_letter(i + 1)
                
                # Set column width
                column_width = max(
                    df_processed[column].astype(str).map(len).max(),  # Width based on cell content
                    len(str(column))                                  # Width based on header
                )
                # Add some padding
                column_width = min(column_width + 2, 50)    # Cap width at 50
                worksheet.column_dimensions[col_letter].width = column_width
                
                # Apply numeric format if applicable
                if column in numeric_formats:
                    for row in range(2, len(df_processed) + 2):  # +2 because we have header and 1-indexing
                        cell = worksheet.cell(row=row, column=i+1)
                        if numeric_formats[column] == '0.00%':
                            cell.number_format = numbers.FORMAT_PERCENTAGE_00
                        elif numeric_formats[column] == '$#,##0.00':
                            cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
                        elif numeric_formats[column] == '#,##0.00':
                            cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                        elif numeric_formats[column] == '0.00':
                            cell.number_format = numbers.FORMAT_NUMBER_00
                        elif numeric_formats[column] == 'yyyy-mm-dd':
                            cell.number_format = numbers.FORMAT_DATE_YYYYMMDD
                
            # Apply borders to all cells
            for row in range(2, len(df_processed) + 2):  # +2 because we have header and 1-indexing
                for col_num in range(1, len(df_processed.columns) + 1):
                    cell = worksheet.cell(row=row, column=col_num)
                    if not cell.border:  # Don't override if we've already set a border
                        cell.border = Border(
                            bottom=Side(border_style='thin'),
                            top=Side(border_style='thin'),
                            left=Side(border_style='thin'),
                            right=Side(border_style='thin')
                        )
        
        buffer.seek(0)
        b64 = base64.b64encode(buffer.read()).decode()
        href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="dataset.xlsx">Download Excel</a>'
        return href
    elif file_format == "json":
        json_str = df.to_json(orient="records")
        b64 = base64.b64encode(json_str.encode()).decode()
        href = f'<a href="data:file/json;base64,{b64}" download="dataset.json">Download JSON</a>'
        return href
    else:
        return ""


def process_files(files: List[Any], prompt: str) -> Tuple[Optional[pd.DataFrame], List[Dict[str, Any]]]:
    """
    Process uploaded files, extract text, and generate structured data.
    
    Args:
        files: List of uploaded files
        prompt: User's natural language prompt
        
    Returns:
        Tuple containing pandas DataFrame and structured content
    """
    if not files:
        st.warning("Please upload at least one file.")
        return None, []
    
    with st.spinner("Extracting text from files..."):
        all_text = ""
        warnings = []
        document_info = {}
        all_structured_content = []
        
        progress_bar = st.progress(0)
        for i, file in enumerate(files):
            progress_text = st.empty()
            progress_text.text(f"Processing {file.name}...")
            
            file_ext = Path(file.name).suffix.lower()
            
            if file_ext == ".pdf":
                text, failed_pages, doc_info, structured_content = extract_text_from_pdf(file)
                all_text += text
                all_structured_content.extend(structured_content)
                if failed_pages:
                    warnings.append(f"Failed to extract text from pages {', '.join(failed_pages)} in {file.name}")
                if doc_info and not document_info:
                    document_info = doc_info
            
            elif file_ext == ".docx":
                text, doc_info, structured_content = extract_text_from_docx(file)
                all_text += text
                all_structured_content.extend(structured_content)
                if doc_info and not document_info:
                    document_info = doc_info
            
            elif file_ext in [".jpg", ".jpeg", ".png"]:
                text, doc_info, structured_content = extract_text_from_image(file)
                all_text += text
                all_structured_content.extend(structured_content)
                if doc_info and not document_info:
                    document_info = doc_info
            
            else:
                warnings.append(f"Unsupported file type: {file.name}")
            
            progress_bar.progress((i + 1) / len(files))
        
        progress_bar.progress(100)
        
        # Display any warnings
        for warning in warnings:
            st.warning(warning)
        
        # Display document type if detected
        if document_info and document_info.get("document_type") != "unknown":
            st.info(f"📄 Detected document type: **{document_info.get('document_type')}** (confidence: {document_info.get('confidence', 0):.2f})")
            
            # Show a sample of extracted structured content
            if len(all_structured_content) > 0:
                with st.expander("View extracted document structure"):
                    # Show a sample of questions, tables, etc.
                    questions = [item for item in all_structured_content if item.get("type") == "question"]
                    tables = [item for item in all_structured_content if item.get("type") == "table"]
                    formulas = [item for item in all_structured_content if item.get("type") in ["formula", "math_block"]]
                    
                    if questions:
                        st.subheader(f"Extracted {len(questions)} questions")
                        st.json(questions[0] if questions else {})
                        
                    if tables:
                        st.subheader(f"Extracted {len(tables)} tables")
                        if "table_data" in tables[0] and tables[0]["table_data"]:
                            st.dataframe(sanitize_df(pd.DataFrame(tables[0]["table_data"])))
                            
                    if formulas:
                        st.subheader(f"Extracted {len(formulas)} formulas/equations")
                        st.json(formulas[0] if formulas else {})
    
    if not all_text.strip():
        st.error("No text could be extracted from the uploaded files.")
        return None, []
    
    # Call LLM to process the text
    with st.spinner("Processing with AI. This might take a while depending on the amount of text..."):
        structured_data = call_llm(prompt, all_text, document_info)
    
    if not structured_data:
        st.error("Failed to generate structured data.")
        return None, all_structured_content
    
    # Convert to DataFrame
    try:
        df = pd.DataFrame(structured_data)
        return df, all_structured_content
    except Exception as e:
        st.error(f"Failed to convert data to table format: {e}")
        return None, all_structured_content


def main():
    """Main application function."""
    # File upload section
    st.subheader("1. Upload Files")
    uploaded_files = st.file_uploader(
        "Upload documents (PDF, DOCX, JPG, PNG)",
        type=["pdf", "docx", "jpg", "jpeg", "png"],
        accept_multiple_files=True
    )
    
    # Prompt input section
    st.subheader("2. Describe Your Dataset")
    user_prompt = st.text_area(
        "Describe the structured data you want to extract from these documents:",
        height=100,
        placeholder="Example: Extract a table with columns for name, email, phone, and address from these customer documents."
    )
    
    # Advanced options
    show_advanced = st.checkbox("Show advanced options")
    if show_advanced:
        extract_mode = st.radio(
            "Extraction mode:",
            ["Standard (tabular data)", "Structured (questions, formulas, tables)"],
            help="Standard mode extracts data into regular tables. Structured mode preserves document structure like questions and formulas."
        )
    else:
        extract_mode = "Standard (tabular data)"
    
    # Process button
    if st.button("Process Files", type="primary", disabled=not (uploaded_files and user_prompt)):
        # Process files and generate DataFrame
        result_df, structured_content = process_files(uploaded_files, user_prompt)
        
        if result_df is not None:
            # Store the DataFrame in session state
            st.session_state["result_df"] = result_df
            st.session_state["processing_complete"] = True
            st.session_state["structured_content"] = structured_content
            st.session_state["extract_mode"] = extract_mode
    
    # Display results if available
    if "processing_complete" in st.session_state and st.session_state["processing_complete"]:
        df = st.session_state["result_df"]
        structured_content = st.session_state.get("structured_content", [])
        extract_mode = st.session_state.get("extract_mode", "Standard (tabular data)")
        
        st.subheader("3. Results")
        
        # Display based on extraction mode
        if extract_mode == "Standard (tabular data)":
            st.subheader("Extracted Table")
            st.dataframe(sanitize_df(df))
            
            st.subheader("4. Download Data")
            col1, col2, col3 = st.columns(3)
            
            with col1:
                st.markdown(get_table_download_link(df, "csv"), unsafe_allow_html=True)
            
            with col2:
                st.markdown(get_table_download_link(df, "excel"), unsafe_allow_html=True)
            
            with col3:
                st.markdown(get_table_download_link(df, "json"), unsafe_allow_html=True)
                
        else:
            # Display structured content in tabs
            tab1, tab2 = st.tabs(["Tabular Data", "Structured Content"])
            
            with tab1:
                st.dataframe(sanitize_df(df))
                
                st.subheader("Download Tabular Data")
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    st.markdown(get_table_download_link(df, "csv"), unsafe_allow_html=True)
                
                with col2:
                    st.markdown(get_table_download_link(df, "excel"), unsafe_allow_html=True)
                
                with col3:
                    st.markdown(get_table_download_link(df, "json"), unsafe_allow_html=True)
            
            with tab2:
                st.subheader("Structured Document Content")
                
                # Group by content type
                questions = [item for item in structured_content if item.get("type") == "question"]
                tables = [item for item in structured_content if item.get("type") == "table"]
                formulas = [item for item in structured_content if item.get("type") in ["formula", "math_block"]]
                images = [item for item in structured_content if item.get("type") == "image"]
                
                if questions:
                    with st.expander(f"Questions ({len(questions)})", expanded=True):
                        for i, question in enumerate(questions[:5]):  # Show first 5 questions
                            st.markdown(f"#### Question {i+1}")
                            st.write(question.get("content", ""))
                            
                            # Show options if available
                            if "options" in question:
                                st.markdown("**Options:**")
                                for letter, text in question["options"].items():
                                    st.markdown(f"**{letter})** {text}")
                            
                            # Show answer if available
                            if "answer" in question:
                                st.markdown(f"**Answer:** {', '.join(question['answer'])}")
                            
                            # Show snippets if available
                            if "snippets" in question and question["snippets"]:
                                for snippet in question["snippets"]:
                                    snippet_path = os.path.join(tempfile.gettempdir(), 
                                                              f"pdf2dataset_{int(time.time())}", snippet)
                                    if os.path.exists(snippet_path):
                                        st.image(snippet_path, caption=f"Question {i+1} image")
                
                if tables:
                    with st.expander(f"Tables ({len(tables)})", expanded=True):
                        for i, table in enumerate(tables[:3]):  # Show first 3 tables
                            st.markdown(f"#### Table {i+1}")
                            if "table_data" in table and table["table_data"]:
                                st.dataframe(sanitize_df(pd.DataFrame(table["table_data"])))
                            else:
                                st.write(table.get("content", ""))
                
                if formulas:
                    with st.expander(f"Formulas ({len(formulas)})", expanded=True):
                        for i, formula in enumerate(formulas[:5]):  # Show first 5 formulas
                            st.markdown(f"#### Formula {i+1}")
                            st.latex(formula.get("content", ""))
                            
                            # Show snippets if available
                            if "snippets" in formula and formula["snippets"]:
                                for snippet in formula["snippets"]:
                                    snippet_path = os.path.join(tempfile.gettempdir(), 
                                                              f"pdf2dataset_{int(time.time())}", snippet)
                                    if os.path.exists(snippet_path):
                                        st.image(snippet_path, caption=f"Formula {i+1} image")
                
                # Download full structured content
                st.download_button(
                    label="Download Full Structured Content (JSON)",
                    data=json.dumps(structured_content, indent=2),
                    file_name="structured_content.json",
                    mime="application/json"
                )


if __name__ == "__main__":
    main() 